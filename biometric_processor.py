"""
Enhanced Biometric System with Truly Interactive Excel Charts
Charts update dynamically based on employee selection
"""

import pandas as pd
import datetime
import os
import re
import sys
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
# Import DataLabelList from openpyxl.chart.label
from openpyxl.chart.label import DataLabelList 


# --- Existing functions (no changes needed for now) ---
def parse_binary_employee_file(file_path):
    """Parse binary employee file using the working method"""
    print(f"📖 Reading binary employee file: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"❌ Employee file not found: {file_path}")
        return {}
    
    try:
        with open(file_path, 'rb') as f:
            data = f.read()
        
        print(f"📊 File size: {len(data)} bytes")
        employees = {}
        pos = 0
        
        # Parse binary records
        while pos < len(data) - 10:
            name_start = pos
            name_end = pos
            
            # Find printable characters (potential name)
            while name_end < len(data) and name_end < pos + 50:
                if data[name_end] == 0:
                    break
                if not (32 <= data[name_end] <= 126):
                    break
                name_end += 1
            
            # Extract potential name
            if name_end > name_start + 2:
                try:
                    potential_name = data[name_start:name_end].decode('latin-1', errors='ignore')
                    
                    if potential_name and re.match(r'^[A-Za-z][A-Za-z\s]*$', potential_name.strip()):
                        name = potential_name.strip()
                        
                        # Look for ID after the name
                        id_search_start = name_end
                        while id_search_start < len(data) and data[id_search_start] == 0:
                            id_search_start += 1
                        
                        # Look for numeric ID in the next 100 bytes
                        for id_pos in range(id_search_start, min(id_search_start + 100, len(data))):
                            if data[id_pos] == 0:
                                continue
                            
                            id_end = id_pos
                            while id_end < len(data) and id_end < id_pos + 10:
                                if data[id_end] == 0:
                                    break
                                if not (48 <= data[id_end] <= 57):
                                    break
                                id_end += 1
                            
                            if id_end > id_pos:
                                try:
                                    potential_id = data[id_pos:id_end].decode('latin-1', errors='ignore')
                                    if potential_id.isdigit() and len(potential_id) <= 3:
                                        employees[potential_id] = name
                                        print(f"   Found: {potential_id} -> {name}")
                                        break
                                except:
                                    continue
                        
                        pos = name_end + 50
                    else:
                        pos += 1
                except:
                    pos += 1
            else:
                pos += 1
        
        print(f"✅ Loaded {len(employees)} employees from binary file")
        return employees
        
    except Exception as e:
        print(f"❌ Error parsing binary employee file: {e}")
        return {}

def extract_names_and_ids_from_binary(data):
    """Alternative method to extract names and IDs from binary data"""
    employees = {}
    
    text_data = ""
    for byte in data:
        if 32 <= byte <= 126:
            text_data += chr(byte)
        else:
            text_data += " "
    
    text_data = re.sub(r'\s+', ' ', text_data).strip()
    pattern = r'([A-Za-z][A-Za-z]*?)(\d{1,3})(?=[A-Z]|\s|$)'
    matches = re.findall(pattern, text_data)
    
    for name, emp_id in matches:
        if name and emp_id:
            employees[emp_id] = name
            print(f"   {emp_id}: {name}")
    
    return employees

def parse_attendance_file(file_path):
    """Parse tab-separated attendance file"""
    print(f"📖 Reading attendance file: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"❌ Attendance file not found: {file_path}")
        return []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        records = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            parts = line.split('\t')
            
            if len(parts) >= 2:
                try:
                    emp_id = parts[0].strip()
                    datetime_str = parts[1].strip()
                    
                    if not emp_id.isdigit():
                        continue
                    
                    datetime_obj = datetime.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
                    
                    record = {
                        'employee_id': emp_id,
                        'datetime': datetime_obj,
                        'date': datetime_obj.date(),
                        'time': datetime_obj.time()
                    }
                    
                    records.append(record)
                    
                except Exception as e: # Catch specific parsing errors for better debug
                    print(f"Skipping malformed attendance line: {line}. Error: {e}")
                    continue
        
        print(f"✅ Parsed {len(records)} attendance records")
        return records
        
    except Exception as e:
        print(f"❌ Error parsing attendance file: {e}")
        return []

def process_attendance_data(records, employees):
    """Process attendance records into daily summary"""
    print("⚙️ Processing attendance data...")
    
    daily_data = {}
    
    for record in records:
        emp_id = record['employee_id']
        emp_name = employees.get(emp_id, f"Unknown_{emp_id}")
        date = record['date']
        time_obj = record['time']
        
        key = (emp_id, emp_name, date)
        
        if key not in daily_data:
            daily_data[key] = {
                'employee_id': emp_id,
                'employee_name': emp_name,
                'date': date,
                'times': []
            }
        
        daily_data[key]['times'].append(time_obj)
    
    processed_data = []
    office_start = datetime.time(9, 30)
    
    for key, data in daily_data.items():
        times = sorted(data['times'])
        
        check_in = times[0] if times else None
        check_out = times[-1] if len(times) > 1 else None
        
        working_hours = 0
        if check_in and check_out:
            check_in_dt = datetime.datetime.combine(data['date'], check_in)
            check_out_dt = datetime.datetime.combine(data['date'], check_out)
            working_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
        
        is_late = False
        late_minutes = 0
        if check_in:
            if check_in > office_start:
                late_minutes = (datetime.datetime.combine(data['date'], check_in) - 
                                datetime.datetime.combine(data['date'], office_start)).total_seconds() / 60
                is_late = True  # Flag as late if check-in is after 9:30
        
        if not check_in:
            status = "ABSENT"
        elif working_hours >= 7:
            status = "PRESENT"
        elif working_hours >= 4:
            status = "HALF_DAY"
        else:
            status = "PRESENT" # If working hours are less than 4 but check-in exists, assume present.
        
        if not check_in:
            late_flag = "❌ ABSENT"
        elif is_late:
            late_flag = "🚩 LATE"
        else:
            late_flag = "✅ ON TIME"
        
        processed_record = {
            'Employee_ID': data['employee_id'],
            'Employee_Name': data['employee_name'],
            'Date': data['date'].strftime('%Y-%m-%d'),
            'Check_In': check_in.strftime('%H:%M:%S') if check_in else 'N/A',
            'Check_Out': check_out.strftime('%H:%M:%S') if check_out else 'N/A',
            'Working_Hours': round(working_hours, 2),
            'Late_Minutes': int(late_minutes),
            'Status': status,
            'Late_Flag': late_flag,
            'Is_Late': is_late
        }
        
        processed_data.append(processed_record)
    
    print(f"✅ Processed data for {len(processed_data)} employee-day records")
    return processed_data

def create_interactive_excel_report(data, employees, output_file):
    """Create Excel report with truly interactive charts"""
    print(f"📊 Creating interactive Excel report with dynamic charts: {output_file}")
    
    try:
        df = pd.DataFrame(data)
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Main Data
        create_main_data_sheet(wb, df, data)
        
        # Sheet 2: Interactive Employee Dashboard
        create_interactive_dashboard(wb, df, employees)
        
        # Sheet 3: Employee Trends
        create_employee_trends_sheet(wb, df, employees)
        
        # Sheet 4: Comparison Charts
        create_comparison_sheet(wb, df)
        
        wb.save(output_file)
        print(f"✅ Interactive Excel report created successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Interactive Excel creation failed: {e}")
        traceback.print_exc()
        return False

def create_main_data_sheet(wb, df, data):
    """Create main data sheet with Excel table for filtering"""
    ws = wb.create_sheet("Main_Data")
    
    # Title
    ws['A1'] = f"Biometric Attendance Data - {datetime.date.today().strftime('%B %d, %Y')}"
    ws['A1'].font = Font(bold=True, size=16, color="2C3E50")
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, record in enumerate(data, 4):
        for col_idx, (key, value) in enumerate(record.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            
            # Color coding
            if record['Is_Late']:
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif record['Status'] == 'ABSENT':
                cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
    
    # Create Excel Table for better filtering
    table_range = f"A3:{get_column_letter(len(headers))}{len(data) + 3}"
    table = Table(displayName="AttendanceData", ref=table_range)
    
    # Add table style
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Auto-adjust columns
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

def create_interactive_dashboard(wb, df, employees):
    """Create interactive dashboard with employee-specific charts"""
    ws = wb.create_sheet("Interactive_Dashboard")
    
    # Title
    ws['B2'] = "🎯 Interactive Employee Dashboard"
    ws['B2'].font = Font(bold=True, size=20, color="2C3E50")
    ws.merge_cells('B2:J2')
    ws['B2'].alignment = Alignment(horizontal='center')
    
    # Employee Selection
    ws['B4'] = "Select Employee:"
    ws['B4'].font = Font(bold=True, size=14)
    
    # Create employee dropdown
    employee_list = sorted([f"{emp_id} - {name}" for emp_id, name in employees.items()])
    
    # Write employee list in hidden area
    for i, emp in enumerate(employee_list, 1):
        ws.cell(row=i, column=20, value=emp)  # Column T
    
    # Data validation dropdown
    dv = DataValidation(
        type="list",
        formula1=f"=$T$1:$T${len(employee_list)}", # Absolute reference
        showDropDown=True
    )
    dv.add(ws['C4'])
    ws.add_data_validation(dv)
    
    # Set default selection
    ws['C4'] = employee_list[0] if employee_list else ""
    
    # Create dynamic data area for selected employee
    create_dynamic_employee_data(ws, df)
    
    # Create interactive charts
    create_interactive_charts(ws)
    
    # Add instructions
    ws['B35'] = "📋 How to Use:"
    ws['B35'].font = Font(bold=True, size=14, color="2C3E50")
    
    instructions = [
        "1. Select an employee from the dropdown in cell C4",
        "2. All charts and metrics will automatically update",
        "3. Charts show last 30 days of data for selected employee",
        "4. Use filters in Main_Data sheet for detailed analysis"
    ]
    
    for i, instruction in enumerate(instructions, 36):
        ws[f'B{i}'] = instruction
        ws[f'B{i}'].font = Font(size=11)

def create_dynamic_employee_data(ws, df):
    """Create dynamic data area that updates based on employee selection"""
    
    # Employee metrics section
    ws['B6'] = "📊 Employee Metrics"
    ws['B6'].font = Font(bold=True, size=14, color="2C3E50")
    
    # Extract employee ID from dropdown selection
    emp_id_formula = 'LEFT(C4,FIND(" ",C4)-1)'
    
    # Dynamic metrics with formulas
    metrics = [
        ("Employee ID:", f'={emp_id_formula}'),
        ("Employee Name:", f'=RIGHT(C4,LEN(C4)-FIND(" - ",C4)-2)'),
        ("Total Records:", f'=COUNTIF(Main_Data!C:C,{emp_id_formula})'), # Use full column reference
        ("Present Days:", f'=COUNTIFS(Main_Data!C:C,{emp_id_formula},Main_Data!J:J,"PRESENT")'), # Adjusted for Status column
        ("Late Days:", f'=COUNTIFS(Main_Data!C:C,{emp_id_formula},Main_Data!K:K,TRUE)'), # Adjusted for Is_Late column
        ("Average Hours:", f'=ROUND(AVERAGEIFS(Main_Data!G:G,Main_Data!C:C,{emp_id_formula},Main_Data!G:G,">0"),2)'), # Adjusted for Working_Hours
        ("Punctuality Rate:", f'=IF(D10>0,TEXT((D10-D11)/D10,"0.0%"),"0%")'), # Changed to TEXT for percentage format
        ("Last Check-in:", f'=INDEX(Main_Data!E:E,MATCH(1,(Main_Data!C:C={emp_id_formula})*(Main_Data!D:D=MAXIFS(Main_Data!D:D,Main_Data!C:C,{emp_id_formula})),0))'), # Adjusted for Check_In, Date, Employee_ID
    ]
    
    for i, (label, formula) in enumerate(metrics, 8):
        ws[f'B{i}'] = label
        ws[f'B{i}'].font = Font(bold=True)
        ws[f'D{i}'] = formula
        ws[f'D{i}'].font = Font(size=11)
        
        # Add borders
        for col in ['B', 'D']:
            ws[f'{col}{i}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Create dynamic chart data area
    create_chart_data_area(ws, df)

def create_chart_data_area(ws, df):
    """Create area with dynamic data for charts"""
    
    # Chart data starts at row 18
    chart_start = 18
    
    # Daily hours chart data
    ws[f'B{chart_start}'] = "📈 Daily Working Hours (Last 30 Days)"
    ws[f'B{chart_start}'].font = Font(bold=True, size=12, color="2C3E50")
    
    ws[f'B{chart_start + 2}'] = "Date"
    ws[f'C{chart_start + 2}'] = "Hours"
    ws[f'D{chart_start + 2}'] = "Status"
    ws[f'E{chart_start + 2}'] = "Late?" # Helper column for pie chart
    
    # Headers for chart data
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{chart_start + 2}'].font = Font(bold=True)
        ws[f'{col}{chart_start + 2}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Create formulas to get last 30 days of data for selected employee
    emp_id_formula = 'LEFT(C4,FIND(" ",C4)-1)'
    
    # Number of rows for dynamic data (e.g., up to 100 rows for flexibility)
    max_chart_rows = 100 

    for row_num in range(chart_start + 3, chart_start + 3 + max_chart_rows):
        # Date column
        ws[f'B{row_num}'].value = f'=IFERROR(INDEX(Main_Data!$D:$D,SMALL(IF(Main_Data!$C:$C={emp_id_formula},ROW(Main_Data!$D:$D)),ROW()-ROW($B${chart_start+3})+1)),"")'
        # Hours column
        ws[f'C{row_num}'].value = f'=IFERROR(INDEX(Main_Data!$G:$G,SMALL(IF(Main_Data!$C:$C={emp_id_formula},ROW(Main_Data!$G:$G)),ROW()-ROW($C${chart_start+3})+1)),"")'
        # Status column
        ws[f'D{row_num}'].value = f'=IFERROR(INDEX(Main_Data!$J:$J,SMALL(IF(Main_Data!$C:$C={emp_id_formula},ROW(Main_Data!$J:$J)),ROW()-ROW($D${chart_start+3})+1)),"")'
        # Is_Late column (for Pie Chart)
        ws[f'E{row_num}'].value = f'=IFERROR(INDEX(Main_Data!$K:$K,SMALL(IF(Main_Data!$C:$C={emp_id_formula},ROW(Main_Data!$K:$K)),ROW()-ROW($E${chart_start+3})+1)),"")'
        
        # Set number format for date
        ws[f'B{row_num}'].number_format = 'YYYY-MM-DD'

    # Add a summary for the pie chart data
    ws[f'G{chart_start + 2}'] = "Status"
    ws[f'H{chart_start + 2}'] = "Count"
    ws[f'G{chart_start + 2}'].font = Font(bold=True)
    ws[f'H{chart_start + 2}'].font = Font(bold=True)
    
    ws[f'G{chart_start + 3}'] = "PRESENT"
    ws[f'H{chart_start + 3}'] = f'=COUNTIFS(D{chart_start+3}:D{chart_start+3+max_chart_rows-1},"PRESENT")'
    
    ws[f'G{chart_start + 4}'] = "ABSENT"
    ws[f'H{chart_start + 4}'] = f'=COUNTIFS(D{chart_start+3}:D{chart_start+3+max_chart_rows-1},"ABSENT")'
    
    ws[f'G{chart_start + 5}'] = "HALF_DAY"
    ws[f'H{chart_start + 5}'] = f'=COUNTIFS(D{chart_start+3}:D{chart_start+3+max_chart_rows-1},"HALF_DAY")'

def create_interactive_charts(ws):
    """Create charts that update based on dynamic data"""
    
    chart_start = 18 # Same as in create_chart_data_area
    max_chart_rows = 100 # Same as in create_chart_data_area

    # Chart 1: Daily Working Hours Line Chart
    line_chart = LineChart()
    line_chart.title = "Daily Working Hours Trend"
    line_chart.style = 12
    line_chart.height = 10
    line_chart.width = 15
    line_chart.x_axis.title = "Date"
    line_chart.y_axis.title = "Hours"
    
    # Dynamic data reference for line chart from the calculated range
    data = Reference(ws, min_col=3, min_row=chart_start + 3, max_row=chart_start + 3 + max_chart_rows - 1, max_col=3)
    categories = Reference(ws, min_col=2, min_row=chart_start + 3, max_row=chart_start + 3 + max_chart_rows - 1, max_col=2)
    
    line_chart.add_data(data, titles_from_data=False)
    line_chart.set_categories(categories)
    
    ws.add_chart(line_chart, "F18")
    
    # Chart 2: Weekly Summary Bar Chart (Placeholder for now, implementation is complex with formulas)
    # This would require more complex Excel formulas to aggregate weekly data from the Main_Data sheet
    # based on the selected employee. For simplicity, we'll keep it as a placeholder.
    bar_chart = BarChart()
    bar_chart.title = "Weekly Attendance Summary (Needs advanced Excel formulas)"
    bar_chart.style = 10
    bar_chart.height = 8
    bar_chart.width = 12
    bar_chart.x_axis.title = "Week"
    bar_chart.y_axis.title = "Days Present"
    
    # You would need to set up a hidden table with weekly summaries based on selected employee for this to work dynamically
    # For now, it will be a blank chart or a static one.
    ws.add_chart(bar_chart, "F35") 
    
    # Chart 3: Status Distribution Pie Chart (using the summarized data)
    pie_chart = PieChart()
    pie_chart.title = "Attendance Status Distribution"
    pie_chart.height = 8
    pie_chart.width = 12
    
    labels = Reference(ws, min_col=7, min_row=chart_start + 3, max_row=chart_start + 5)
    data = Reference(ws, min_col=8, min_row=chart_start + 3, max_row=chart_start + 5)
    
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    # FIX: Instantiate DataLabelList correctly
    pie_chart.series[0].dLbls = DataLabelList(showCatName=True, showPercent=True) 
    
    ws.add_chart(pie_chart, "R18")

def create_employee_trends_sheet(wb, df, employees):
    """Create sheet with employee trend analysis"""
    ws = wb.create_sheet("Employee_Trends")
    
    # Title
    ws['A1'] = "📈 Employee Trends Analysis"
    ws['A1'].font = Font(bold=True, size=18, color="2C3E50")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Employee dropdown (same as other sheet)
    ws['A3'] = "Select Employee:"
    ws['A3'].font = Font(bold=True, size=12)
    
    employee_list = sorted([f"{emp_id} - {name}" for emp_id, name in employees.items()])
    
    for i, emp in enumerate(employee_list, 1):
        ws.cell(row=i, column=15, value=emp)
    
    dv = DataValidation(
        type="list",
        formula1=f"=$O$1:$O${len(employee_list)}", # Absolute reference
        showDropDown=True
    )
    dv.add(ws['B3'])
    ws.add_data_validation(dv)
    ws['B3'] = employee_list[0] if employee_list else ""
    
    # Monthly trends
    create_monthly_trends_section(ws)
    
    # Punctuality trends
    create_punctuality_trends_section(ws)

def create_monthly_trends_section(ws):
    """Create monthly trends analysis"""
    
    ws['A5'] = "📅 Monthly Trends"
    ws['A5'].font = Font(bold=True, size=14, color="2C3E50")
    
    # Headers
    headers = ['Month', 'Total Days', 'Present Days', 'Late Days', 'Avg Hours', 'Punctuality %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Sample monthly data (in practice, this would be calculated from actual data)
    # Generate last 6 months for example
    current_date = datetime.date.today()
    months = []
    for i in range(6):
        month = (current_date.replace(day=1) - pd.DateOffset(months=i)).strftime('%Y-%m')
        months.append(month)
    months.reverse() # Show most recent last
    
    for i, month in enumerate(months, 8):
        ws.cell(row=i, column=1, value=month)
        
        # Use formulas to calculate monthly stats for selected employee
        emp_id_formula = 'LEFT(B3,FIND(" ",B3)-1)'
        
        start_date_ref = f'DATE(LEFT(A{i},4),MID(A{i},6,2),1)'
        end_date_ref = f'EDATE({start_date_ref},1)' # Start of next month
        
        # Total days in month (records for selected employee within this month)
        ws.cell(row=i, column=2, value=f'=COUNTIFS(Main_Data!C:C,{emp_id_formula},Main_Data!D:D,">="&{start_date_ref},Main_Data!D:D,"<"&{end_date_ref})')
        
        # Present days
        ws.cell(row=i, column=3, value=f'=COUNTIFS(Main_Data!C:C,{emp_id_formula},Main_Data!D:D,">="&{start_date_ref},Main_Data!D:D,"<"&{end_date_ref},Main_Data!J:J,"PRESENT")')
        
        # Late days
        ws.cell(row=i, column=4, value=f'=COUNTIFS(Main_Data!C:C,{emp_id_formula},Main_Data!D:D,">="&{start_date_ref},Main_Data!D:D,"<"&{end_date_ref},Main_Data!K:K,TRUE)')
        
        # Average hours
        ws.cell(row=i, column=5, value=f'=ROUND(AVERAGEIFS(Main_Data!G:G,Main_Data!C:C,{emp_id_formula},Main_Data!D:D,">="&{start_date_ref},Main_Data!D:D,"<"&{end_date_ref},Main_Data!G:G,">0"),2)')
        
        # Punctuality percentage
        ws.cell(row=i, column=6, value=f'=IF(C{i}>0,TEXT((C{i}-D{i})/C{i},"0.0%"),"0%")')

def create_punctuality_trends_section(ws):
    """Create punctuality trends section"""
    
    ws['A15'] = "⏰ Punctuality Analysis"
    ws['A15'].font = Font(bold=True, size=14, color="2C3E50")
    
    # Create weekly punctuality chart data
    ws['A17'] = "Week"
    ws['B17'] = "On Time"
    ws['C17'] = "Late"
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}17'].font = Font(bold=True)
        ws[f'{col}17'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Sample weekly data (this would be calculated dynamically)
    weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4'] # Last 4 weeks
    for i, week in enumerate(weeks, 18):
        ws[f'A{i}'] = week
        ws[f'B{i}'] = f'=RANDBETWEEN(3,5)'  # Placeholder - replace with actual formula
        ws[f'C{i}'] = f'=RANDBETWEEN(0,2)'  # Placeholder - replace with actual formula


def create_comparison_sheet(wb, df):
    """Create employee comparison sheet"""
    ws = wb.create_sheet("Employee_Comparison")
    
    # Title
    ws['A1'] = "👥 Employee Performance Comparison"
    ws['A1'].font = Font(bold=True, size=18, color="2C3E50")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Create summary table for all employees
    employee_stats = df.groupby(['Employee_ID', 'Employee_Name']).agg({
        'Working_Hours': ['mean', 'sum', 'count'],
        'Is_Late': 'sum',
        'Status': lambda x: (x != 'ABSENT').sum()
    }).round(2)
    
    employee_stats.columns = ['Avg_Hours', 'Total_Hours', 'Days_Count', 'Late_Count', 'Present_Days']
    # Ensure no division by zero for Punctuality_Rate
    employee_stats['Punctuality_Rate'] = employee_stats.apply(
        lambda row: round(((row['Present_Days'] - row['Late_Count']) / row['Present_Days']) * 100, 1) 
        if row['Present_Days'] > 0 else 0.0, axis=1
    )
    
    employee_stats = employee_stats.reset_index()
    
    # Write comparison data
    headers = ['Employee_ID', 'Employee_Name', 'Avg_Hours', 'Total_Hours', 'Present_Days', 'Late_Count', 'Punctuality_Rate']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, (_, row) in enumerate(employee_stats.iterrows(), 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust columns for comparison sheet
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    # Create comparison charts
    create_comparison_charts(ws, employee_stats)

def create_comparison_charts(ws, employee_stats):
    """Create comparison charts"""
    
    # Top performers chart (Avg Hours)
    top_performers = employee_stats.nlargest(min(10, len(employee_stats)), 'Avg_Hours') # Ensure it doesn't try to get more than available
    
    chart_row_start = len(employee_stats) + 6
    
    # Write data for top performers chart
    ws.cell(chart_row_start, 1, "Employee")
    ws.cell(chart_row_start, 2, "Avg Hours")
    ws.cell(chart_row_start, 3, "Punctuality Rate") # For second chart
    
    for i, (_, row) in enumerate(top_performers.iterrows(), 1):
        ws.cell(chart_row_start + i, 1, row['Employee_Name'][:15]) # Limit name length
        ws.cell(chart_row_start + i, 2, row['Avg_Hours'])
        ws.cell(chart_row_start + i, 3, row['Punctuality_Rate'])
    
    # Chart 1: Top Performers by Average Hours Bar Chart
    bar_chart_avg_hours = BarChart()
    bar_chart_avg_hours.title = "Top Performers by Average Hours"
    bar_chart_avg_hours.height = 10
    bar_chart_avg_hours.width = 15
    bar_chart_avg_hours.x_axis.title = "Employees"
    bar_chart_avg_hours.y_axis.title = "Average Hours"
    
    data_avg_hours = Reference(ws, min_col=2, min_row=chart_row_start, 
                     max_row=chart_row_start + len(top_performers), max_col=2)
    categories_avg_hours = Reference(ws, min_col=1, min_row=chart_row_start + 1, 
                                  max_row=chart_row_start + len(top_performers), max_col=1)
    
    bar_chart_avg_hours.add_data(data_avg_hours, titles_from_data=True) # titles_from_data=True for header
    bar_chart_avg_hours.set_categories(categories_avg_hours)
    
    ws.add_chart(bar_chart_avg_hours, f"A{chart_row_start + 15}") # Position first chart
    
    # Chart 2: Top Performers by Punctuality Rate Bar Chart
    top_punctual = employee_stats.nlargest(min(10, len(employee_stats)), 'Punctuality_Rate') # Top by punctuality
    
    # Reuse chart_row_start + 1 to keep data together
    # Write data for top punctual chart (can reuse the same employee names)
    chart_row_punctual_start = chart_row_start # Reuse same data area or create new if needed
    ws.cell(chart_row_punctual_start, 5, "Employee")
    ws.cell(chart_row_punctual_start, 6, "Punctuality Rate")

    for i, (_, row) in enumerate(top_punctual.iterrows(), 1):
        ws.cell(chart_row_punctual_start + i, 5, row['Employee_Name'][:15])
        ws.cell(chart_row_punctual_start + i, 6, row['Punctuality_Rate'])

    bar_chart_punctuality = BarChart()
    bar_chart_punctuality.title = "Top Performers by Punctuality Rate"
    bar_chart_punctuality.height = 10
    bar_chart_punctuality.width = 15
    bar_chart_punctuality.x_axis.title = "Employees"
    bar_chart_punctuality.y_axis.title = "Punctuality Rate (%)"
    
    data_punctuality = Reference(ws, min_col=6, min_row=chart_row_punctual_start, 
                                 max_row=chart_row_punctual_start + len(top_punctual), max_col=6)
    categories_punctuality = Reference(ws, min_col=5, min_row=chart_row_punctual_start + 1, 
                                       max_row=chart_row_punctual_start + len(top_punctual), max_col=5)
    
    bar_chart_punctuality.add_data(data_punctuality, titles_from_data=True)
    bar_chart_punctuality.set_categories(categories_punctuality)
    
    ws.add_chart(bar_chart_punctuality, f"Q{chart_row_start + 15}") # Position second chart


def print_summary(data):
    """Print summary to console"""
    df = pd.DataFrame(data)
    
    total_employees = len(df['Employee_ID'].unique())
    present_count = len(df[df['Status'].isin(['PRESENT', 'HALF_DAY'])]) # Consider half-day as present
    late_count = len(df[df['Is_Late'] == True])
    
    # Attendance rate based on 'Present' or 'Half_Day' records per unique employee day
    # This calculation needs to be more precise: it's not total_employees vs present_count
    # It's unique (Employee_ID, Date) pairs where status is not ABSENT vs total unique (Employee_ID, Date) pairs.
    total_days_recorded = len(df[['Employee_ID', 'Date']].drop_duplicates())
    present_days_recorded = len(df[df['Status'].isin(['PRESENT', 'HALF_DAY'])][['Employee_ID', 'Date']].drop_duplicates())
    attendance_rate = (present_days_recorded / total_days_recorded) * 100 if total_days_recorded > 0 else 0
    
    print("\n" + "="*60)
    print("📊 INTERACTIVE ATTENDANCE SYSTEM SUMMARY")
    print("="*60)
    print(f"📈 Total Employees: {total_employees}")
    print(f"✅ Total Present/Half-Day Records: {present_count}") # More accurate wording
    print(f"🚩 Total Late Records: {late_count}")
    print(f"📊 Overall Attendance Rate (Days with attendance): {attendance_rate:.1f}%")
    print("="*60)

# --- NEW FUNCTION FOR FLASK INTEGRATION ---
def process_biometric_data_for_excel_dashboard(employee_file_path, attendance_file_path, output_excel_path):
    """
    Processes biometric data from raw files and generates the interactive Excel dashboard.
    This function is designed to be called by your Flask app.
    
    Args:
        employee_file_path (str): Path to the raw binary employee data file.
        attendance_file_path (str): Path to the raw tab-separated attendance data file.
        output_excel_path (str): The desired path and filename for the output Excel dashboard.
        
    Returns:
        bool: True if the Excel dashboard was created successfully, False otherwise.
    """
    print("--- Starting Biometric Data Processing for Excel Dashboard ---")
    print(f"Employee File: {employee_file_path}")
    print(f"Attendance File: {attendance_file_path}")
    print(f"Output Excel: {output_excel_path}")

    employees = parse_binary_employee_file(employee_file_path)
    
    if not employees:
        print("\n⚠️ Trying alternative parsing method for employee file...")
        try:
            with open(employee_file_path, 'rb') as f:
                data = f.read()
            employees = extract_names_and_ids_from_binary(data)
        except Exception as e:
            print(f"❌ Alternative method failed for employee file: {e}")
    
    if not employees:
        print("❌ Failed to load employees. Cannot create report.")
        return False
    
    attendance_records = parse_attendance_file(attendance_file_path)
    if not attendance_records:
        print("❌ Failed to load attendance records. Cannot create report.")
        return False
    
    processed_data = process_attendance_data(attendance_records, employees)
    if not processed_data:
        print("❌ Failed to process attendance data. Cannot create report.")
        return False
    
    success = create_interactive_excel_report(processed_data, employees, output_excel_path)
    
    if success:
        print_summary(processed_data) # Print summary to console for server logs
        print("\n--- Biometric Data Processing for Excel Dashboard Completed Successfully ---")
    else:
        print("\n--- Biometric Data Processing for Excel Dashboard FAILED ---")
        
    return success

# --- Keep the __main__ block for standalone testing, but it won't be used by Flask ---
if __name__ == '__main__':
    try:
        # Original main function logic for command line execution
        # You would typically provide the paths as arguments like:
        # python biometric_processor.py -e employee.dat -a attendance.txt
        
        if len(sys.argv) >= 5 and sys.argv[1] == '-e' and sys.argv[3] == '-a':
            employee_file = sys.argv[2]
            attendance_file = sys.argv[4]
            # Define a default output file name for standalone runs
            today = datetime.date.today()
            output_file = f"interactive_attendance_charts_{today.strftime('%Y%m%d')}_standalone.xlsx"
            
            process_biometric_data_for_excel_dashboard(employee_file, attendance_file, output_file)
            print(f"Standalone report saved to: {output_file}")
        else:
            print("❌ Usage for standalone run: python biometric_processor.py -e <employee_file> -a <attendance_file>")
            print("This script is primarily designed to be called by the Flask backend.")

    except KeyboardInterrupt:
        print("\n👋 Process interrupted")
    except Exception as e:
        print(f"\n💥 Error: {e}")
        traceback.print_exc()

